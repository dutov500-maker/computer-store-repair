import json
import os
import io
import base64
import boto3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


CATALOG = [
    # ECO
    {"category": "ECO (Эконом)", "name": "Eco 1 — Ryzen 5 5500 + RTX 1660 Super", "price": 55000},
    {"category": "ECO (Эконом)", "name": "Eco 2 — Intel Core i3 12100F + RTX 2060 Super", "price": 59500},
    {"category": "ECO (Эконом)", "name": "Eco 3 — Intel Core i3 12100F + RTX 3050", "price": 63000},
    {"category": "ECO (Эконом)", "name": "Eco 4 — Intel Core i3 12100F + RX 6600", "price": 66000},
    {"category": "ECO (Эконом)", "name": "Eco 5 — Intel Core i5 12400F + RTX 3060", "price": 70000},
    # SPECIAL
    {"category": "SPECIAL (Специальные)", "name": "Special 1 — Ryzen 5 5600 + RTX 5060", "price": 95000},
    {"category": "SPECIAL (Специальные)", "name": "Special 2 — Intel Core i5 12400F + RTX 5060", "price": 115000},
    {"category": "SPECIAL (Специальные)", "name": "Special 3 — Ryzen 5 7500F + RTX 5060", "price": 130000},
    {"category": "SPECIAL (Специальные)", "name": "Special 4 — Ryzen 5 7500F + RTX 5060 Ti", "price": 140000},
    # PREMIUM
    {"category": "PREMIUM (Премиум)", "name": "Premium 1 — Ryzen 7 8700F + RTX 5060 Ti", "price": 145000},
    {"category": "PREMIUM (Премиум)", "name": "Premium 2 — Ryzen 7 8700F + RTX 5070", "price": 165000},
    {"category": "PREMIUM (Премиум)", "name": "Premium 3 — Ryzen 7 7800X3D + RTX 5070", "price": 205000},
    {"category": "PREMIUM (Премиум)", "name": "Premium 4 — Ryzen 7 7800X3D + RTX 5070 Ti", "price": 225000},
    # ULTRA
    {"category": "ULTRA (Ультра)", "name": "Ultra 1 — Ryzen 7 7800X3D + RTX 5070 Ti", "price": 265000},
    {"category": "ULTRA (Ультра)", "name": "Ultra 2 — Ryzen 7 7800X3D + RTX 5080", "price": 299000},
    {"category": "ULTRA (Ультра)", "name": "Ultra 3 — Ryzen 7 7800X3D + RTX 5080", "price": 309000},
    # ELITE
    {"category": "ELITE (Элит)", "name": "Elite 1 — Ryzen 9 9950X3D + RTX 5090", "price": 560000},
]

SERVICES = [
    # Компьютеры
    {"category": "Ремонт компьютеров", "name": "Диагностика компьютера", "price": "Бесплатно"},
    {"category": "Ремонт компьютеров", "name": "Чистка от пыли", "price": "от 500 ₽"},
    {"category": "Ремонт компьютеров", "name": "Замена термопасты", "price": "от 800 ₽"},
    {"category": "Ремонт компьютеров", "name": "Установка Windows", "price": "от 1 000 ₽"},
    {"category": "Ремонт компьютеров", "name": "Ремонт материнской платы", "price": "от 2 000 ₽"},
    {"category": "Ремонт компьютеров", "name": "Замена блока питания", "price": "от 500 ₽"},
    {"category": "Ремонт компьютеров", "name": "Апгрейд компьютера", "price": "от 1 000 ₽"},
    {"category": "Ремонт компьютеров", "name": "Восстановление данных", "price": "от 3 000 ₽"},
    {"category": "Ремонт компьютеров", "name": "Настройка BIOS", "price": "от 800 ₽"},
    {"category": "Ремонт компьютеров", "name": "Удаление вирусов", "price": "от 1 500 ₽"},
    {"category": "Ремонт компьютеров", "name": "Замена кулера", "price": "от 700 ₽"},
    {"category": "Ремонт компьютеров", "name": "Ремонт видеокарты", "price": "от 2 500 ₽"},
    {"category": "Ремонт компьютеров", "name": "Установка драйверов", "price": "от 500 ₽"},
    {"category": "Ремонт компьютеров", "name": "Настройка сети", "price": "от 800 ₽"},
    {"category": "Ремонт компьютеров", "name": "Замена жесткого диска", "price": "от 1 200 ₽"},
    # Ноутбуки
    {"category": "Ремонт ноутбуков", "name": "Ремонт ноутбука", "price": "от 1 000 ₽"},
    {"category": "Ремонт ноутбуков", "name": "Замена экрана ноутбука", "price": "от 3 000 ₽"},
    {"category": "Ремонт ноутбуков", "name": "Чистка ноутбука от пыли", "price": "от 800 ₽"},
    {"category": "Ремонт ноутбуков", "name": "Замена клавиатуры ноутбука", "price": "от 1 500 ₽"},
    # Планшеты
    {"category": "Ремонт планшетов", "name": "Ремонт планшета", "price": "от 1 500 ₽"},
    {"category": "Ремонт планшетов", "name": "Замена экрана планшета", "price": "от 2 500 ₽"},
    # Телефоны
    {"category": "Ремонт телефонов", "name": "Ремонт телефона", "price": "от 1 000 ₽"},
    {"category": "Ремонт телефонов", "name": "Замена экрана телефона", "price": "от 2 000 ₽"},
    {"category": "Ремонт телефонов", "name": "Замена батареи телефона", "price": "от 1 200 ₽"},
    {"category": "Ремонт телефонов", "name": "Ремонт разъема зарядки", "price": "от 1 500 ₽"},
    {"category": "Ремонт телефонов", "name": "Замена камеры телефона", "price": "от 1 800 ₽"},
]


def make_header_style():
    font = Font(bold=True, color="FFFFFF", size=11)
    fill = PatternFill("solid", fgColor="1E3A5F")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return font, fill, alignment


def make_category_style():
    font = Font(bold=True, color="FFFFFF", size=10)
    fill = PatternFill("solid", fgColor="2E5984")
    alignment = Alignment(horizontal="left", vertical="center", indent=1)
    return font, fill, alignment


def thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def build_workbook():
    wb = openpyxl.Workbook()

    # ── Лист 1: Каталог ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Каталог"
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 46
    ws1.column_dimensions["C"].width = 18

    h_font, h_fill, h_align = make_header_style()
    c_font, c_fill, c_align = make_category_style()
    border = thin_border()

    # Заголовок листа
    ws1.merge_cells("A1:C1")
    title_cell = ws1["A1"]
    title_cell.value = "КАТАЛОГ ИГРОВЫХ КОМПЬЮТЕРОВ — КомпЛаб"
    title_cell.font = Font(bold=True, color="FFFFFF", size=13)
    title_cell.fill = PatternFill("solid", fgColor="0F2A47")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    # Шапка таблицы
    headers = ["Категория", "Конфигурация", "Цена, ₽"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col, value=h)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = h_align
        cell.border = border
    ws1.row_dimensions[2].height = 22

    row = 3
    current_cat = None
    for item in CATALOG:
        if item["category"] != current_cat:
            current_cat = item["category"]
            ws1.merge_cells(f"A{row}:C{row}")
            cat_cell = ws1.cell(row=row, column=1, value=current_cat)
            cat_cell.font = c_font
            cat_cell.fill = c_fill
            cat_cell.alignment = c_align
            ws1.row_dimensions[row].height = 20
            row += 1

        ws1.cell(row=row, column=1).border = border
        ws1.cell(row=row, column=2, value=item["name"]).border = border
        ws1.cell(row=row, column=2).alignment = Alignment(vertical="center", indent=1)
        price_cell = ws1.cell(row=row, column=3, value=item["price"])
        price_cell.number_format = '#,##0 ₽'
        price_cell.alignment = Alignment(horizontal="right", vertical="center")
        price_cell.border = border
        ws1.row_dimensions[row].height = 18
        # Чередование строк
        if row % 2 == 0:
            for c in range(1, 4):
                ws1.cell(row=row, column=c).fill = PatternFill("solid", fgColor="EEF4FB")
        row += 1

    # ── Лист 2: Услуги и ремонт ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Услуги и ремонт")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 46
    ws2.column_dimensions["C"].width = 18

    ws2.merge_cells("A1:C1")
    t2 = ws2["A1"]
    t2.value = "УСЛУГИ И РЕМОНТ — КомпЛаб"
    t2.font = Font(bold=True, color="FFFFFF", size=13)
    t2.fill = PatternFill("solid", fgColor="0F2A47")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    for col, h in enumerate(["Категория", "Услуга", "Стоимость"], 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = h_align
        cell.border = border
    ws2.row_dimensions[2].height = 22

    row = 3
    current_cat = None
    for item in SERVICES:
        if item["category"] != current_cat:
            current_cat = item["category"]
            ws2.merge_cells(f"A{row}:C{row}")
            cat_cell = ws2.cell(row=row, column=1, value=current_cat)
            cat_cell.font = c_font
            cat_cell.fill = c_fill
            cat_cell.alignment = c_align
            ws2.row_dimensions[row].height = 20
            row += 1

        ws2.cell(row=row, column=1).border = border
        ws2.cell(row=row, column=2, value=item["name"]).border = border
        ws2.cell(row=row, column=2).alignment = Alignment(vertical="center", indent=1)
        price_cell = ws2.cell(row=row, column=3, value=item["price"])
        price_cell.alignment = Alignment(horizontal="right", vertical="center")
        price_cell.border = border
        ws2.row_dimensions[row].height = 18
        if row % 2 == 0:
            for c in range(1, 4):
                ws2.cell(row=row, column=c).fill = PatternFill("solid", fgColor="EEF4FB")
        row += 1

    return wb


def handler(event: dict, context) -> dict:
    """Генерирует XLS прайс-лист КомпЛаб и возвращает публичную ссылку на файл в S3."""
    if event.get("httpMethod") == "OPTIONS":
        return {
            "statusCode": 200,
            "headers": {
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "GET, OPTIONS",
                "Access-Control-Allow-Headers": "Content-Type",
                "Access-Control-Max-Age": "86400",
            },
            "body": "",
        }

    wb = build_workbook()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    file_bytes = buf.read()

    s3 = boto3.client(
        "s3",
        endpoint_url="https://bucket.poehali.dev",
        aws_access_key_id=os.environ["AWS_ACCESS_KEY_ID"],
        aws_secret_access_key=os.environ["AWS_SECRET_ACCESS_KEY"],
    )
    key = "files/komplab-pricelist.xlsx"
    s3.put_object(
        Bucket="files",
        Key=key,
        Body=file_bytes,
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ContentDisposition='attachment; filename="komplab-pricelist.xlsx"',
    )

    cdn_url = f"https://cdn.poehali.dev/projects/{os.environ['AWS_ACCESS_KEY_ID']}/bucket/{key}"

    return {
        "statusCode": 200,
        "headers": {
            "Access-Control-Allow-Origin": "*",
            "Content-Type": "application/json",
        },
        "body": json.dumps({"url": cdn_url}),
    }
